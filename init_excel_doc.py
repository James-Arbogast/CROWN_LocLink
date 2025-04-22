from pathlib import Path
import openpyxl as pxl
from util.xliff.xliff import File as XLIFF
from util.xliff.xliff import TransUnit
from util.LanguageCodes import Language

crwn_path = Path(r"M:\Projects\Crown\Translation Docs\CROWN_メインシナリオ_第1部_0303 (1) 2.xlsx")
wkbk = pxl.load_workbook(crwn_path)
sh = wkbk['01']

spkr_col = 'G'
stg_dir_col = 'H'
dialg_col = 'J'
type_col = 'F'

stg_id = 0
dialg_id = 0
spkr_id = 0

inbox = Path(r'M:\Projects\Crown\memoQ Folders\01 INBOX')
xlf = XLIFF.create(Path(r'01.xliff'), 'ja', 'en')

note = ""
for i in range(2, sh.max_row):
    type = sh[f'{type_col}{i}'].value
    spkr = sh[f'{spkr_col}{i}'].value
    stg = sh[f'{stg_dir_col}{i}'].value
    dialg = sh[f'{dialg_col}{i}'].value
    if str(type) != "None":
        note = type
    if str(spkr) != "None":
        spkr = str(spkr)
        new_spkr_tu = TransUnit.create(Language.Japanese,
                                    Language.English,
                                    f'spkr_{spkr_id}',
                                    spkr,
                                    "",
                                    [note],
                                    False,
                                    "needs-translation")
        xlf.trans_units.set_value(new_spkr_tu)
        spkr_id+=1
        
    if str(dialg) != "None":
        dialg = str(dialg)
        while str(sh[f'{dialg_col}{i+1}'].value) != "None" and i < sh.max_row:
            i+=1    
            dialg+=f"\n{sh[f'{dialg_col}{i}'].value}"
        new_dialg_tu = TransUnit.create(Language.Japanese,
                                    Language.English,
                                    f'dialg_{dialg_id}',
                                    dialg,
                                    "",
                                    [note],
                                    False,
                                    "needs-translation")
        xlf.trans_units.set_value(new_dialg_tu)
        dialg_id+=1
        
    if str(stg) != "None":
        stg = str(stg)
        while str(sh[f'{stg_dir_col}{i+1}'].value) != "None" and i < sh.max_row:
            i+=1     
            stg+=f"\n{sh[f'{stg_dir_col}{i}'].value}"
        new_stg_tu = TransUnit.create(Language.Japanese,
                                    Language.English,
                                    f'stg_{stg_id}',
                                    stg,
                                    "",
                                    [note],
                                    False,
                                    "needs-translation")
        xlf.trans_units.set_value(new_stg_tu)
        stg_id+=1       
    xlf.save_in_directory(inbox)
    
    
