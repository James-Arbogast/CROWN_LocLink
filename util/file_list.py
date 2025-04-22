from pathlib import Path
import openpyxl as pxl
from util.xliff.xliff import File as XLIFF
from util.xliff.xliff import TransUnit
from util.LanguageCodes import Language
from util.MemoQDatabase import MemoQDatabase
from util.lxtxt.lxtxtDatabase import LXTXTDatabase
import os
import xlsxwriter
import random
import time
from util.data_tracking.count_JPC import count_JPC

#Random Color Generator
rlite = lambda: random.randint(150,255)

class FileList:
    def __init__(self, filelist_location: Path, memoq_db: MemoQDatabase, lxtxt_db: LXTXTDatabase):
        self.filelist_location = filelist_location
        self.memoq_db = memoq_db
        self.lxtxt_db = lxtxt_db
        self.workbook = None
    
    #Color Formatting Prep.
    def add_format(self, color):
        global format
        format=self.workbook.add_format({
        'bg_color' : color
        })
    
    def find_leaves(self, path):
        for root, dirs, files in os.walk(path):
            if not dirs:
                yield root
      
    ##Longest Path Finding functions
    def nesting(self, path):
        #counts how often os.path.split works on path
        c = 0
        head = tail = path
        while head and tail:
            head, tail = os.path.split(head)
            c += 1
        return c
    
    def longest_path(self, paths):
        return max(paths,key=self.nesting)
        
    def create_file_list(self, fullpath: bool):
        print("Commencing File List")

        ##Start the progress bar.
        numberoffiles = 0
        for root, dirs, files in os.walk(self.memoq_db.input_folder):
            for f in files:
                numberoffiles += 1	
        
        print("Finding longest path")
        longestPath = self.longest_path(self.find_leaves(self.memoq_db.input_folder))
        relLongestPath = longestPath.replace(os.path.abspath(self.memoq_db.input_folder),"")
        maxdirs = relLongestPath.count("\\")

        workbook = xlsxwriter.Workbook(self.filelist_location)
        self.workbook = workbook
        sheet = self.workbook.add_worksheet()
        bold = self.workbook.add_format({'bold': True})

        ##Set up the sheet's columns.
        columnNo = 0
        if fullpath == True:
            sheet.write(0,columnNo,"Full Path",bold)
            columnNo+=1
        curcolumnno = columnNo
        for column in range(curcolumnno,maxdirs):
            sheet.write(0, columnNo,"Path " + str(column),bold)
            columnNo+=1
        sheet.write(0,columnNo,"Filename",bold)
        columnNo+=1
        sheet.write(0,columnNo, "File JPC", bold)

        #Fill the Excel.
        writeRow = 1
        
        print("Generating colors")
        ##Generate a list of every unique directory in the folder, then assign a color to it.
        foldercolor = {}
        for directory in os.walk(self.memoq_db.input_folder):
            dirname = directory[0]
            foldername = dirname[dirname.rfind("\\")+1:]
            if foldername not in foldercolor:
                generatecolor = ('#%02X%02X%02X' % (rlite(),rlite(),rlite()))
                foldercolor[foldername] = generatecolor

        print("Processing files")
        start_time = time.time()
        filesfinished = 0
        for root, dirs, files in os.walk(self.memoq_db.input_folder):
            for filename in files:
                ### get total JPC count of xliff file
                total_jpc = self.get_file_jpc_count(Path(os.path.join(root,filename)))
                print(total_jpc)
                #estimate time remaining
                time_elapsed = time.time() - start_time
                if filesfinished != 0:
                    if round(time_elapsed) % 10 == 0:
                        filesleft = numberoffiles - filesfinished
                        estimatedtime = (time_elapsed / filesfinished) * filesleft
                        timeestimate_str = time.strftime("%M:%S", time.gmtime(estimatedtime))
                #code
                fullPath = os.path.join(root,filename)
                relativepath = (fullPath.replace(os.path.abspath(self.memoq_db.input_folder),""))[1:]
                relativepathnofolder= relativepath.replace(filename,"")
                pathlist = relativepathnofolder.split("\\")
                columnNo = 0
                if fullpath == 1:
                    sheet.write(writeRow,columnNo,fullPath)
                    columnNo += 1
                for part in pathlist:
                    if part in foldercolor:
                        self.add_format(foldercolor[part])
                        sheet.write(writeRow,columnNo,part,format)
                    else:
                        sheet.write(writeRow,columnNo,part)
                    columnNo += 1
                sheet.write(writeRow,maxdirs,filename)
                sheet.write(writeRow, maxdirs+1, total_jpc)
                writeRow += 1
                filesfinished += 1
        self.workbook.close()
        
    def get_file_jpc_count(self, xliff_path: Path):
        total_jpc = 0
        xliff = XLIFF.from_file(xliff_path, self.memoq_db.input_folder)
        for key, t_unit in xliff.trans_units.items():
            if t_unit.source:
                total_jpc += count_JPC(t_unit.source)
        return total_jpc


'''
crwn_path = Path(r"M:\Projects\Crown\Translation Docs\CROWN_メインシナリオ_第1部_0303 (1) 2.xlsx")
wkbk = pxl.load_workbook(crwn_path)
sh = wkbk['01']

spkr_col = 'G'
stg_dir_col = 'H'
dialg_col = 'J'
type_col = 'F'

stg_id = 0
dialg_id = 0
spkr_id = 0

inbox = Path(r'M:\Projects\Crown\memoQ Folders\01 INBOX')
xlf = XLIFF.create(Path(r'01.xliff'), 'ja', 'en')

note = ""
for i in range(2, sh.max_row):
    type = sh[f'{type_col}{i}'].value
    spkr = sh[f'{spkr_col}{i}'].value
    stg = sh[f'{stg_dir_col}{i}'].value
    dialg = sh[f'{dialg_col}{i}'].value
    if str(type) != "None":
        note = type
    if str(spkr) != "None":
        spkr = str(spkr)
        new_spkr_tu = TransUnit.create(Language.Japanese,
                                    Language.English,
                                    f'spkr_{spkr_id}',
                                    spkr,
                                    "",
                                    [note],
                                    False,
                                    "needs-translation")
        xlf.trans_units.set_value(new_spkr_tu)
        spkr_id+=1
        
    if str(dialg) != "None":
        dialg = str(dialg)
        while str(sh[f'{dialg_col}{i+1}'].value) != "None" and i < sh.max_row:
            i+=1    
            dialg+=f"\n{sh[f'{dialg_col}{i}'].value}"
        new_dialg_tu = TransUnit.create(Language.Japanese,
                                    Language.English,
                                    f'dialg_{dialg_id}',
                                    dialg,
                                    "",
                                    [note],
                                    False,
                                    "needs-translation")
        xlf.trans_units.set_value(new_dialg_tu)
        dialg_id+=1
        
    if str(stg) != "None":
        stg = str(stg)
        while str(sh[f'{stg_dir_col}{i+1}'].value) != "None" and i < sh.max_row:
            i+=1     
            stg+=f"\n{sh[f'{stg_dir_col}{i}'].value}"
        new_stg_tu = TransUnit.create(Language.Japanese,
                                    Language.English,
                                    f'stg_{stg_id}',
                                    stg,
                                    "",
                                    [note],
                                    False,
                                    "needs-translation")
        xlf.trans_units.set_value(new_stg_tu)
        stg_id+=1       
    xlf.save_in_directory(inbox)
'''   
    
